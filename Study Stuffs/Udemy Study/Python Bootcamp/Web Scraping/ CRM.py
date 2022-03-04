from email import message
from glob import glob
from tkinter import *
from tkinter import messagebox
from turtle import bgcolor
from matplotlib.pyplot import text, title

from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as WDW
from selenium.common.exceptions import TimeoutException as TE
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import click

from openpyxl import load_workbook
from sqlalchemy import column
from sympy import div

# DRIVER

driver = webdriver.Chrome(
        executable_path=r"/Users/clcx/Documents/Calvin/Random/chromedriver")

# buka web crm
driver.get("https://crm.xcyunxt.com/xccrm/#/clue/index")

# set id pass
usn = '15006106002'
pwd = 'A@1234xc'

# excel workbook
filepath = '/Users/clcx/Documents/Calvin/CRM/Blank.xlsx'
wb = load_workbook(filepath)
i = 2
div = 2

sheetRange = wb['Sheet1']

def open_chrome():
    driver = webdriver.Chrome(
            executable_path=r"/Users/clcx/Documents/Calvin/Random/chromedriver")
    
# FUNCTIONS

def enter_path() -> None:
    global wb
    global sheetRange
    global filepath
    filepath = filepath_box.get()
    wb = load_workbook(filepath)
    sheetRange = wb['Sheet1']
    print(filepath)

def save_excel():
    wb.save(filepath)

def plus_i():
    global i
    i += 1
    i_label = Label(window, text=str(i), font=(14))
    i_label.grid(row=2, column=2)

def minus_i():
    global i
    i -= 1
    i_label = Label(window, text=str(i), font=(14))
    i_label.grid(row=2, column=2)

def plus_div():
    global div
    div += 1
    div_label = Label(window, text='Div : '+str(div), font=(14))
    div_label.grid(row=5, column=2)

def minus_div():
    global div
    div -= 1
    div_label = Label(window, text='Div : '+str(div), font=(14))
    div_label.grid(row=5, column=2)

def clear():  # clear form
    try:
        WDW(driver, 5).until(EC.presence_of_element_located(
            (By.XPATH, '//*[@id="app"]/div/div[2]/section/div/div/div[3]/div/div[2]')))

        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/div[2]/section/div/div/div[3]/div/div[2]/form/div[2]/div/div/div[1]/input').clear()
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/div[2]/section/div/div/div[3]/div/div[2]/form/div[1]/div/div/input').clear()
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/div[2]/section/div/div/div[3]/div/div[2]/form/div[6]/div[1]/div/div/div[1]/input').clear()
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/div[2]/section/div/div/div[3]/div/div[2]/form/div[6]/div[2]/div[1]/div/div/div/input').clear()
    except:
        print('NOTHING TO CLEAR')

def clickable(element: str) -> None:  # clickable
    try:
        element = WDW(driver, 5).until(EC.presence_of_element_located(
            (By.XPATH, element)))
        element.click()
    except:
        pass

def save():
    try:
        element = WDW(driver, 5).until(EC.presence_of_element_located(
            (By.XPATH, '//*[@id="app"]/div/div[2]/section/div/div/div[3]/div/div[2]/form/div[7]/div/button[1]')))
        element.click()
        print(str(i)+' SAVED')
    except:
        pass

def close_form():
    try:
        element = WDW(driver, 5).until(EC.presence_of_element_located(
            (By.XPATH, '//*[@id="app"]/div/div[2]/section/div/div/div[3]/div/div[1]/button/i')))
        element.click()
        print('FORM CLOSED')
    except:
        pass

def close_repetition1():
    try:
        element = WDW(driver, 5).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[2]/div/div[1]/button')))
        element.click()
        print('REPETITION 1 CLOSED')
    except:
        pass

def close_repetition2():
    try:
        element = WDW(driver, 5).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[3]/div/div[1]/button/i')))
        element.click()
        print('REPETITION 2 CLOSED')
    except:
        pass

def close_repetition():
    try:
        for i in range(4):
            close_repetition1()
            close_repetition2()
    except:
        print('NOT SHOWING REPETITION ALERT')

def close_duplicate():
    try:
        WDW(driver, 5).until(EC.presence_of_element_located(
        (By.XPATH, '/html/body/div[4]/div/div[2]/div/div/div[1]')))
        driver.find_element(By.XPATH,'/html/body/div[4]/div/div[1]/button/i').click()
        print('duplicate alert closed')
    except:
        print('NOT SHOWING DUPLICATE ALERT')

def close_existing():
    try:
        WDW(driver, 5).until(EC.presence_of_element_located(
        (By.XPATH, '/html/body/div[3]/div/div[1]')))
        driver.find_element(By.XPATH,'/html/body/div[3]/div/div[1]/button/i').click()
        print('existing alert closed')
    except:
        print('NOT SHOWING EXISTING ALERT')
    time.sleep(2)

# login dll

def start():

    username = driver.find_element(By.NAME, 'username')
    username.clear()
    username.send_keys(usn)
    password = driver.find_element(By.NAME, 'password')
    password.clear()
    password.send_keys(pwd)
    driver.find_element(By.XPATH, '//*[@id="app"]/div/div/form/button').click()

    # pilih CRM
    clickable('//*[@id="app"]/div/div[1]/div[4]/div[1]/div[1]/img')

    # buka bagian input data
    # ganti focus ke tab baru
    driver.switch_to.window(driver.window_handles[1])

    # panah bawah
    clickable('//*[@id="app"]/div/div[1]/div[2]/div[1]/div/ul/div[2]/li/div/span')
    driver.implicitly_wait(3)

    # pilih isi data
    clickable('//*[@id="app"]/div/div[1]/div[2]/div[1]/div/ul/div[2]/li/ul/div[6]/a/li')
    time.sleep(3)

# input data

def input_data(i: int) -> None:
    # tambah data baru
    clickable('//*[@id="app"]/div/div[2]/section/div/div/div[1]/div[1]/button[1]')

    olshop_name = sheetRange['D'+str(i)].value
    firm_name = sheetRange['C'+str(i)].value
    boss_name = sheetRange['B'+str(i)].value
    phone_number = sheetRange['A'+str(i)].value

    # masukin data
    try:
        WDW(driver, 5).until(EC.presence_of_element_located(
            (By.XPATH, '//*[@id="app"]/div/div[2]/section/div/div/div[3]/div/div[2]')))
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/div[2]/section/div/div/div[3]/div/div[2]/form/div[2]/div/div/div[1]/input').send_keys(olshop_name)
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/div[2]/section/div/div/div[3]/div/div[2]/form/div[1]/div/div/input').send_keys(firm_name)
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/div[2]/section/div/div/div[3]/div/div[2]/form/div[6]/div[1]/div/div/div[1]/input').send_keys(boss_name)
        driver.find_element(
            By.XPATH, '//*[@id="app"]/div/div[2]/section/div/div/div[3]/div/div[2]/form/div[6]/div[2]/div[1]/div/div/div/input').send_keys(phone_number)
    except:
        print('FORM IS NOT SHOWING')


def note_dup(i: int) -> None:
    # kalo muncul alert catet ke excel + clear form
    global div
    try:
        WDW(driver, 5).until(EC.presence_of_element_located(
        (By.XPATH, '/html/body/div['+str(div)+']/div/div[2]/div/div/div[1]')))
        sheetRange['E'+str(i)].value = driver.find_element(By.XPATH,'/html/body/div['+str(div)+']/div/div[2]/div/div/div[2]/div/div[3]/span[2]').text
        sheetRange['F'+str(i)].value = driver.find_element(By.XPATH,'/html/body/div['+str(div)+']/div/div[2]/div/div/div[2]/div/div[5]/span[2]').text
        clear()
        wb.save(filepath)
        print('NOTED on '+str(i))
    except:
        print('DID NOT NOTE')
        print('/html/body/div['+str(div)+']/div/div[2]/div/div/div[2]/div/div[3]/span[2]')
        print('/html/body/div['+str(div)+']/div/div[2]/div/div/div[2]/div/div[5]/span[2]')


window = Tk()
window.title('Money Maker')
window.geometry('300x170')

i_label = Label(window, text=str(i), font=(14))
i_label.grid(row=2, column=2)

add_button = Button(window, command=plus_i, text='+', font=(14))
add_button.grid(row=2, column=1, sticky=E)

minus_button = Button(window, command=minus_i, text='-', font=(14))
minus_button.grid(row=2, column=1, sticky=W)

addx_button = Button(window, command=plus_div, text='+', font=(14))
addx_button.grid(row=5, column=1, sticky=E)

div_label = Label(window, text='Div : '+str(div), font=(14))
div_label.grid(row=5, column=2)

minusx_button = Button(window, command=minus_div, text='-', font=(14))
minusx_button.grid(row=5, column=1, sticky=W)

start_button = Button(window, command=start, text='START', font=(14), bg='red')
start_button.grid(row=6, column=2, sticky=E)

save_excel_button = Button(window, command=save_excel, text='SAVE XL', font=(14), bg='red')
save_excel_button.grid(row=6, column=1, sticky=W)

input_button = Button(window, command=lambda: input_data(i), text='INPUT', font=(14))
input_button.grid(row=3, column=1, sticky=W)

note_button = Button(window, command=lambda: note_dup(i), text='NOTE', font=(14))
note_button.grid(row=4, column=1, sticky=W)

save_button = Button(window, command=save, text='SAVE', font=(14))
save_button.grid(row=4, column=1, sticky=E)

clear_button = Button(window, command=clear, text='CLEAR', font=(14))
clear_button.grid(row=3, column=1, sticky=E)

filepath_box = Entry(window, textvariable=filepath, font=(14))
filepath_box.grid(row=1, column=1)

enter_path_button = Button(window, command=enter_path, text='ENTER', font=(14))
enter_path_button.grid(row=1, column=2, sticky=W)

window.mainloop()
